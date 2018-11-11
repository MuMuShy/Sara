#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import speech_recognition as sr
import tempfile
from gtts import gTTS
from pygame import mixer
import requests
from time import ctime
import time
import pyttsx3
import imutils
from bs4 import BeautifulSoup
import tkinter as tk
import os
from PIL import Image
from PIL import ImageTk
from pprint import pprint
import re
from SaraModules import *
#Sara指令表
def ReadAllCommand():
    from openpyxl import load_workbook
    wb=load_workbook('Sara指令表/Sara指令表.xlsx')
    ws=wb.get_active_sheet()
    SaraCommandDict=dict()
    for i in range(2,ws.max_column+1):
        SaraCommandDict[ws.cell(row=1,column=i).value]=[ws.cell(row=3,column=i).value,ws.cell(row=4,column=i).value]
    return SaraCommandDict
#Sara的喇賽聊天機器模組
class SaraChat():
    def getChatterbot(self):
        from chatterbot import ChatBot
        chatbot=ChatBot('Sara',trainer = 'chatterbot.trainers.ChatterBotCorpusTrainer')
        # 載入(簡體)中文的問候語言庫
        chatbot.train("chatterbot.corpus.chinese.greetings")
        # 載入(簡體)中文的對話語言庫
        chatbot.train("chatterbot.corpus.chinese.conversations")
        return chatbot
#Sara 的本體
def CreatGUI():
    #讀取所有指令檔
    SaraCommandDict=ReadAllCommand()
    #載入Chatbot
    SaraChatterBot=SaraChat()
    SaraChatterBot=SaraChatterBot.getChatterbot()
    def SearchCommand(talk):
        talk=talk.lower()
        print(talk)
        for t in SaraCommandDict.keys():
            if re.findall(t,talk):
                SpeakChinese('好的 請稍等呦')
                time.sleep(2)
                #執行excel下的指令
                exec(SaraCommandDict[t][0])
                #講出callback
                SpeakChinese(SaraCommandDict[t][1])
                time.sleep(2)
                return
        #沒有指令 喇賽
        data=SaraChatterBot.get_response(talk)
        str(data)
        if data:
            SpeakChinese(data)
        print(data)
    #sara的耳朵
    def Listen():
        print('sara is listening...')
        r=sr.Recognizer()
        with sr.Microphone() as source:
            r.adjust_for_ambient_noise(source)
            audio=r.listen(source)
        data=""
        #開始語音辨識
        try:
            data=r.recognize_google(audio,language='zh-TW')
            print("you said"+data)
        except sr.UnknownValueError:
            print("Google Speech Recognition could not understand audio")
            SetIText('無法辨識 請再說一遍')
        except sr.RequestError as e:
            print("Could not request results from Google Speech Recognition service; {0}".format(e))
        return data
    #Sara說英文
    def callback():
        print('a')
    def Speak(sentence):
        mixer.init()
        with tempfile.NamedTemporaryFile(delete=True)as fp:
            tts=gTTS(text=sentence,lang='en-US')
            tts.save("{}.mp3".format(fp.name))
            mixer.music.load("{}.mp3".format(fp.name))
            mixer.music.play()
            SetSaraText(sentence)
    #Sara 中文
    def SpeakChinese(sentence):
        mixer.init()
        with tempfile.NamedTemporaryFile(delete=True)as fp:
            tts=gTTS(text=sentence,lang='zh-TW')
            tts.save("{}.mp3".format(fp.name))
            mixer.music.load("{}.mp3".format(fp.name))
            mixer.music.play()
            SetSaraText(sentence)
    def Speak2(sentence):
        mixer.init()
        engine = mixer.init()
        rate = engine.getProperty('rate')
        engine.setProperty('rate', rate)
        voices = engine.getProperty('voices')
        engine.setProperty('voice',voices[7].id)
        engine.say(sentence)
        engine.runAndWait()
    #Sara的大腦
    def Sara(data):
        SearchCommand(data)
    #當視窗被關注 被調用(問候)
    def Onfocus(event):
        first=False
        if first==False:
            if event.widget==window:
                print('i am focus on')
                window.update()
                Speak('hi i am Sara')
                time.sleep(1)
                Speak('What can i help you?')
                time.sleep(1)
                SetSaraText('Hi i am Sara What Can I help You')
                time.sleep(1)
                first=True
    #sara 使用者文字輸入接口
    def SaraByText():
        command=UserEntry.get()        
    def SetSaraText(saratext):
        SaraCanvas.delete('SaraText')
        SaraCanvas.create_text((10,20),text=saratext,anchor="w",fill="white",tag='SaraText')
        SaraCanvas.update()
    def test():
        print('button test')
    def SaraMainLoop():
        if(mixer.music.get_busy()):
            print('')
        window.after(1000, SaraMainLoop)
    #退出程序
    def on_closing():
        window.destroy()
        os._exit(0)
    #創建Tk window
    window=tk.Tk()
    window.title('Sara')
    window.geometry('400x400')
    #創建canvas並設置背景圖
    canvas=tk.Canvas(window,height=200,width=400)
    img = Image.open("sara.jpg")
    img = img.resize((400,200), Image.ANTIALIAS)
    photoImg = ImageTk.PhotoImage(img)
    canvas.create_image(0,0, image=photoImg,anchor='nw')
    master=window
    canvas.pack()
    #使用者文本輸入
    UserEntry=tk.Entry(window)
    UserEntry.pack(padx=100,pady=0)
    #使用者輸入按鈕
    UserInputbutton=tk.Button(window,text="輸入",width=10,height=2,command=lambda:SearchCommand(UserEntry.get()))
    UserInputbutton.pack(padx=0,pady=0)
    #使用者說話按鈕
    UserSaybutton=tk.Button(window,text='說話',width=10,height=2,command=lambda:Sara(Listen()))
    UserSaybutton.pack(padx=0,pady=20)
    #創建Sara回話Canvas
    SaraCanvas=tk.Canvas(window,height=300,width=400,bg="black")
    SaraCanvas.create_text((10,20),text="Sara: hi i am sara ha ha",anchor="w",fill="white",tag='SaraText')
    SaraCanvas.pack()
    #綁定event
    window.bind("<FocusIn>", Onfocus)
    window.protocol("WM_DELETE_WINDOW", on_closing)
    #讓window 在最前視窗
    window.lift()
    window.attributes('-topmost', True)
    os.system('''/usr/bin/osascript -e 'tell app "Finder" to set frontmost of process "Python" to true' ''')
    #window loop
    window.after(1000,SaraMainLoop)
    window.mainloop()
def main():
    CreatGUI()
if __name__ == '__main__':
    main()


# In[ ]:





# In[ ]:




